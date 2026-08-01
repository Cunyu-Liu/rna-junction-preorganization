#!/usr/bin/env python3
"""Audit Denny workbook semantics using aggregate counts only.

The output intentionally omits sequence strings, individual labels, effect
values, and row-level identifiers. It is a provenance/semantics audit, not a
label export.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

import openpyxl


TARGET_COUNTS = (1687, 1713, 1636)
MARKERS = {
    "measured": re.compile(r"measured", re.IGNORECASE),
    "interpolated": re.compile(r"interpol", re.IGNORECASE),
    "raw": re.compile(r"\braw\b", re.IGNORECASE),
    "censor": re.compile(r"censor|lower.?bound|limit", re.IGNORECASE),
    "replicate": re.compile(r"replic", re.IGNORECASE),
    "bootstrap": re.compile(r"bootstrap", re.IGNORECASE),
    "covariance": re.compile(r"covar|variance", re.IGNORECASE),
    "unit_kcal": re.compile(r"kcal|kilocal", re.IGNORECASE),
    "nine_bp": re.compile(r"\b9\s*(?:bp|bps|base.?pair)", re.IGNORECASE),
    "eleven_bp": re.compile(r"\b11\s*(?:bp|bps|base.?pair)", re.IGNORECASE),
}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def header_text(value: object) -> str:
    if value is None:
        return "<blank>"
    text = str(value).strip()
    if not text:
        return "<blank>"
    # Headers are not row-level data, but keep the output bounded and safe.
    return text[:120]


def numeric_value(value: object) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else None
    return None


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--expected-sha256", required=True)
    args = parser.parse_args()

    observed_sha256 = sha256(args.input)
    if observed_sha256 != args.expected_sha256:
        print(
            json.dumps(
                {
                    "status": "BLOCKED_HASH_MISMATCH",
                    "expected_sha256": args.expected_sha256,
                    "observed_sha256": observed_sha256,
                },
                sort_keys=True,
            )
        )
        return 2

    workbook = openpyxl.load_workbook(args.input, read_only=True, data_only=False)
    sheets: list[dict[str, object]] = []
    global_marker_counts: Counter[str] = Counter()
    global_target_count_occurrences: Counter[str] = Counter()
    global_minus_7_1_occurrences = 0
    global_formula_cells = 0

    for worksheet in workbook.worksheets:
        rows = worksheet.iter_rows(values_only=True)
        header_row = None
        header = None
        column_stats: list[dict[str, object]] = []
        header_marker_presence: dict[str, bool] = {name: False for name in MARKERS}
        nonempty_data_rows = 0
        target_rows: Counter[str] = Counter()

        for row_index, row in enumerate(rows, start=1):
            if header is None:
                string_values = [value for value in row if isinstance(value, str) and value.strip()]
                if len(string_values) < 2:
                    continue
                header_row = row_index
                header = row
                headers = [header_text(value) for value in header]
                header_text_blob = " | ".join(headers)
                header_marker_presence = {
                    name: bool(pattern.search(header_text_blob))
                    for name, pattern in MARKERS.items()
                }
                column_stats = [
                    {
                        "header": field,
                        "nonempty_count": 0,
                        "numeric_count": 0,
                        "text_count": 0,
                        "formula_count": 0,
                        "exact_minus_7_1_count": 0,
                        "numeric_less_than_minus_7_1_count": 0,
                        "numeric_equal_to_minus_7_1_count": 0,
                        "numeric_greater_than_minus_7_1_count": 0,
                        "nonnegative_integer_sum": 0,
                        "_nonnegative_integer_values": set(),
                        "_distinct_value_keys": set(),
                        "_distinct_integer_values": set(),
                        "target_count_occurrences": {str(target): 0 for target in TARGET_COUNTS},
                        "marker_counts": {name: 0 for name in MARKERS},
                    }
                    for field in headers
                ]
                continue
            row_nonempty = False
            row_numeric_values: set[int] = set()
            for column_index, value in enumerate(row):
                if column_index >= len(column_stats) or value is None or not str(value).strip():
                    continue
                row_nonempty = True
                stats = column_stats[column_index]
                stats["nonempty_count"] += 1
                stats["_distinct_value_keys"].add((type(value).__name__, str(value)))
                number = numeric_value(value)
                if number is not None:
                    stats["numeric_count"] += 1
                    if number.is_integer():
                        stats["_distinct_integer_values"].add(int(number))
                    if number == -7.1:
                        stats["exact_minus_7_1_count"] += 1
                        stats["numeric_equal_to_minus_7_1_count"] += 1
                        global_minus_7_1_occurrences += 1
                    elif number < -7.1:
                        stats["numeric_less_than_minus_7_1_count"] += 1
                    else:
                        stats["numeric_greater_than_minus_7_1_count"] += 1
                    if (
                        header is not None
                        and column_index < len(header)
                        and str(header[column_index]).strip().lower() == "number of variants"
                        and number.is_integer()
                        and number >= 0
                    ):
                        stats["nonnegative_integer_sum"] += int(number)
                        stats["_nonnegative_integer_values"].add(int(number))
                    if number.is_integer() and int(number) in TARGET_COUNTS:
                        target = str(int(number))
                        stats["target_count_occurrences"][target] += 1
                        global_target_count_occurrences[target] += 1
                        row_numeric_values.add(int(number))
                else:
                    stats["text_count"] += 1
                    text = str(value)
                    if text.startswith("="):
                        stats["formula_count"] += 1
                        global_formula_cells += 1
                    for name, pattern in MARKERS.items():
                        if pattern.search(text):
                            stats["marker_counts"][name] += 1
                            global_marker_counts[name] += 1
            if row_nonempty:
                nonempty_data_rows += 1
                for target in row_numeric_values:
                    target_rows[str(target)] += 1

        for stats in column_stats:
            stats["distinct_value_count"] = len(stats.pop("_distinct_value_keys"))
            stats["distinct_integer_value_count"] = len(stats.pop("_distinct_integer_values"))
            if stats["header"] == "Number of variants":
                stats["nonnegative_integer_values"] = sorted(stats.pop("_nonnegative_integer_values"))
            else:
                stats.pop("_nonnegative_integer_values")

        sheets.append(
            {
                "name": worksheet.title,
                "header_row": header_row,
                "header_marker_presence": header_marker_presence,
                "max_row_reported": worksheet.max_row,
                "max_column_reported": worksheet.max_column,
                "nonempty_data_row_count": nonempty_data_rows,
                "row_count_matches_contract_target": {
                    str(target): nonempty_data_rows == target for target in TARGET_COUNTS
                },
                "target_count_rows_with_numeric_occurrence": dict(target_rows),
                "columns": column_stats,
            }
        )
    workbook.close()

    result = {
        "status": "SEMANTICS_AGGREGATE_AUDIT_COMPLETE",
        "audited_at_utc": datetime.now(timezone.utc).isoformat(),
        "input": str(args.input),
        "output": str(args.output),
        "sha256": observed_sha256,
        "size_bytes": args.input.stat().st_size,
        "target_counts_checked": list(TARGET_COUNTS),
        "target_count_occurrences_across_cells": dict(global_target_count_occurrences),
        "minus_7_1_numeric_occurrences_across_cells": global_minus_7_1_occurrences,
        "marker_counts_across_cells": dict(global_marker_counts),
        "formula_cell_count": global_formula_cells,
        "sheets": sheets,
        "interpretation": {
            "exact_count_reconciliation": "NOT_YET_ESTABLISHED",
            "censor_direction": "NOT_YET_ESTABLISHED",
            "raw_vs_interpolated_semantics": "NOT_YET_ESTABLISHED",
            "replicate_bootstrap_covariance_semantics": "NOT_YET_ESTABLISHED",
            "no_primary_labels_admitted": True,
            "scientific_gate_effect": "NO_PHASE_0_PASS",
        },
        "manual_review_required": [
            "map aggregate counts to the article's named library subsets",
            "confirm the workbook header row and source documentation semantics",
            "confirm whether -7.1 is a censoring boundary and establish its direction",
            "confirm whether replicate/error fields are independent, bootstrap-derived, or covariance-aware",
        ],
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(
        json.dumps(
            {
                "status": result["status"],
                "sha256": observed_sha256,
                "sheet_count": len(sheets),
                "minus_7_1_occurrences": global_minus_7_1_occurrences,
                "target_count_occurrences": dict(global_target_count_occurrences),
            },
            sort_keys=True,
        )
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
