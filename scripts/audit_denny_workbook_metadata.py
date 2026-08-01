#!/usr/bin/env python3
"""Audit Denny supplementary workbook structure without emitting raw values."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl


NUCLEOTIDE_LIKE = re.compile(r"^[ACGUTNRYKMSWBDHVX&./_\-0-9]+$", re.IGNORECASE)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def safe_field(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    # Never emit a sequence-like or unusually long cell into the metadata file.
    if len(text) > 40 or NUCLEOTIDE_LIKE.fullmatch(text.replace(" ", "")):
        return "<masked>"
    return text[:80]


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--expected-sha256", required=True)
    args = parser.parse_args()

    observed_sha256 = sha256(args.input)
    if observed_sha256 != args.expected_sha256:
        print(
            json.dumps(
                {
                    "status": "BLOCKED_HASH_MISMATCH",
                    "expected_sha256": args.expected_sha256,
                    "observed_sha256": observed_sha256,
                },
                sort_keys=True,
            )
        )
        return 2

    workbook = openpyxl.load_workbook(args.input, read_only=True, data_only=False)
    sheets: list[dict[str, object]] = []
    for worksheet in workbook.worksheets:
        nonempty_rows = 0
        nonempty_cells = 0
        first_nonempty_row = None
        candidate_headers: list[str] | None = None
        candidate_header_row = None
        for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            values = [value for value in row if value is not None and str(value).strip()]
            if not values:
                continue
            nonempty_rows += 1
            nonempty_cells += len(values)
            if first_nonempty_row is None:
                first_nonempty_row = row_index
            if candidate_headers is None:
                string_values = [safe_field(value) for value in values if isinstance(value, str)]
                string_values = [value for value in string_values if value is not None]
                if len(string_values) >= 2:
                    candidate_headers = string_values
                    candidate_header_row = row_index

        sheets.append(
            {
                "name": worksheet.title,
                "max_row_reported": worksheet.max_row,
                "max_column_reported": worksheet.max_column,
                "nonempty_row_count": nonempty_rows,
                "nonempty_cell_count": nonempty_cells,
                "first_nonempty_row": first_nonempty_row,
                "candidate_header_row": candidate_header_row,
                "candidate_header_fields": candidate_headers,
            }
        )
    workbook.close()

    field_text = json.dumps(sheets, ensure_ascii=False).lower()
    required_terms = {
        term: term in field_text
        for term in ("raw", "interpol", "censor", "covar", "replic", "delta", "kcal")
    }
    result = {
        "status": "METADATA_AUDIT_COMPLETE",
        "audited_at_utc": datetime.now(timezone.utc).isoformat(),
        "input": str(args.input),
        "output": str(args.output),
        "sha256": observed_sha256,
        "size_bytes": args.input.stat().st_size,
        "sheet_count": len(sheets),
        "sheets": sheets,
        "required_term_presence_in_candidate_headers": required_terms,
        "raw_values_emitted": False,
        "scientific_gate_effect": "NO_PHASE_0_PASS",
        "manual_review_required": [
            "reconcile workbook counts against contract and article claims",
            "confirm raw/interpolated/censor direction semantics from documentation",
            "confirm replicate/bootstrap/covariance fields and units",
            "build traceable motif matching table before using any labels",
        ],
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({"status": result["status"], "sha256": observed_sha256, "sheet_count": len(sheets)}, sort_keys=True))
    return 0


if __name__ == "__main__":
    sys.exit(main())
