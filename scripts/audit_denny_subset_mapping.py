#!/usr/bin/env python3
"""Audit Denny workbook subset counts using aggregate metadata only.

The workbook contains row-level sequence and thermodynamic content. This
script deliberately emits only named sublibrary counts and counts of a
candidate numeric identifier column; it never writes rows, sequences, labels,
or effect values to the output.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import openpyxl


EXPECTED_TARGETS = (1687, 1713, 1636)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def integer(value: object) -> int | None:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return None
    number = float(value)
    if not math.isfinite(number) or not number.is_integer():
        return None
    return int(number)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--expected-sha256", required=True)
    args = parser.parse_args()

    observed = sha256(args.input)
    if observed != args.expected_sha256:
        print(json.dumps({"status": "BLOCKED_HASH_MISMATCH", "expected": args.expected_sha256, "observed": observed}, sort_keys=True))
        return 2

    workbook = openpyxl.load_workbook(args.input, read_only=True, data_only=True)
    descriptions = workbook["sublibrary_descriptions"]
    description_rows: list[dict[str, object]] = []
    for row in descriptions.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        count = integer(row[2]) if len(row) > 2 else None
        description_rows.append({"sublibrary": str(row[0]).strip(), "number_of_variants": count})

    annotations = workbook["library_annotations"]
    candidate_ids: set[int] = set()
    ids_by_annotation: dict[str, set[int]] = defaultdict(set)
    measured_ids: set[int] = set()
    interpolated_ids: set[int] = set()
    rows_with_candidate_id = 0
    for row in annotations.iter_rows(min_row=3, values_only=True):
        candidate = integer(row[9] if len(row) > 9 else None)
        if candidate is None:
            continue
        candidate_ids.add(candidate)
        rows_with_candidate_id += 1
        annotation = str(row[1]).strip() if len(row) > 1 and row[1] is not None else "<blank>"
        ids_by_annotation[annotation].add(candidate)
        if len(row) > 15 and row[15] is not None:
            measured_ids.add(candidate)
        if len(row) > 23 and row[23] is not None:
            interpolated_ids.add(candidate)
    workbook.close()

    group_summary = [
        {"annotation": key, "distinct_candidate_id_count": len(value)}
        for key, value in sorted(ids_by_annotation.items())
    ]
    observed_targets = {
        str(target): target in candidate_ids for target in EXPECTED_TARGETS
    }
    result = {
        "schema_version": "phase0-denny-subset-mapping-v1",
        "status": "SUBSET_MAPPING_AGGREGATE_AUDIT_COMPLETE",
        "audited_at_utc": datetime.now(timezone.utc).isoformat(),
        "input": str(args.input),
        "sha256": observed,
        "sublibrary_descriptions": description_rows,
        "explicit_variant_count_sum": sum(
            row["number_of_variants"] or 0 for row in description_rows
        ),
        "candidate_identifier_column": "library_annotations_column_10_one_based",
        "rows_with_candidate_identifier": rows_with_candidate_id,
        "distinct_candidate_identifier_count": len(candidate_ids),
        "distinct_candidate_identifier_count_by_annotation": group_summary,
        "distinct_candidate_identifier_count_with_measured_nonempty": len(measured_ids),
        "distinct_candidate_identifier_count_with_interpolated_nonempty": len(interpolated_ids),
        "contract_target_counts_observed_as_candidate_identifiers": observed_targets,
        "contract_target_mapping_status": "UNRESOLVED_1687_AND_1636_NOT_ESTABLISHED_BY_THIS_FIELD",
        "raw_sequence_content_emitted": False,
        "primary_labels_admitted": False,
        "scientific_gate_effect": "NO_PHASE_0_PASS",
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({"status": result["status"], "sha256": observed, "distinct_candidate_identifier_count": len(candidate_ids), "explicit_variant_count_sum": result["explicit_variant_count_sum"]}, sort_keys=True))
    return 0


if __name__ == "__main__":
    sys.exit(main())
